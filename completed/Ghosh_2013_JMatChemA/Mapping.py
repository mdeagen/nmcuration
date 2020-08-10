#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 10 15:03:34 2020

@author: davidjany
"""
import os ,shutil
from openpyxl import load_workbook
import pandas as pd

original_path = os.getcwd()

#%% Load spreadsheet with data for all samples

xl = pd.ExcelFile(original_path+'/Samples_10.1039_C3TA10381A.xlsx')
df  = xl.parse(sheet_name=0)




#%% Auxiliary functions

def create_dir(name): #creates a directory and change the location to it
    path = os.getcwd()
    os.mkdir(name)
    os.chdir(path+'/'+name)


def get_master(sample):
    pst = df['pst'][sample]
    if pst =='none':
        master = 'master_template_S1.xlsx'
    elif pst =='polydimethyl-siloxane':
        master = 'master_template.xlsx'
    elif pst =='3-Aminopropyltrimethoxysilane':
        master = 'master_template_APS.xlsx'
    else:
        master = 'master_template_APS_aa.xlsx'
    return master



#%% Mapping
    
def map_data_origin(workbook , sample): #fill in cells in '1. Data Origin' sheet for a sample
    sheet = workbook['1. Data Origin']
    sheet["B5"] = df['sample no.'][sample]
    
def map_material_type(workbook ,sample):  #fill in cells in '2. Material Types' sheet for a sample
    sheet = workbook['2. Material Types']
    sheet["C46"] = df['filler mass fraction'][sample]


def map_thermal(workbook ,sample): #fill in cells in '5.4 Properties-Thermal' sheet for a sample
    sheet = workbook['5.4 Properties-Thermal']
    sheet["C18"] = df['Tc'][sample]
    sheet["D18"] = 'Celsius'
    sheet["C30"] = df['Tm'][sample]
    sheet["D30"] = 'Celsius'
    sheet["C26"] = df['Tg'][sample]
    sheet["D26"] = 'Celsius'
    

def map_rheology(workbook,sample):
    wb = load_workbook(original_path+"/G'(T).xlsx")
    sheet = wb['Feuil1']
    sheet['B2'] = df["G'(30)"][sample]
    sheet['B3'] = df["G'(150)"][sample]
    sheet['B4'] = df["G'(200)"][sample]
    filename = df['sample no.'][sample]+'/'+df['sample no.'][sample]+'_storage_modulus.xlsx' # added sample prefix to filename (M.D.)
    wb.save(filename=filename)
    sheet = workbook['5.6 Properties-Rheological']
    sheet['C14'] = df['sample no.'][sample]+'_storage_modulus.xlsx' # added sample prefix to filename (M.D.)

def copy_image(sample): 
    file = str(df['image microstructure file'][sample])
    if file[-1] =='g':
        shutil.copyfile(original_path+"/Images/"+df['image microstructure file'][sample], 'S'+str(sample+1)+"/"+df['image microstructure file'][sample])      
    else:
        pass
    
def mapping():    #Fill in every cell for every sample
    create_dir('SUBMISSION')
    nb_sample = df['sample no.'].shape[0]
    for i in range(nb_sample):
        sample = df['sample no.'][i]
        os.mkdir(sample) 
        master = get_master(i)
        file = original_path+'/'+master
        workbook = load_workbook(filename=file) 
        
        map_data_origin(workbook ,i)
        map_material_type(workbook ,i)
        map_thermal(workbook ,i)
        map_rheology(workbook ,i)
        
        filename = sample+'/'+sample+'_template.xlsx'
        workbook.save(filename=filename)
        
        
#%% Launch program    
mapping()
    