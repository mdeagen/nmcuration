#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Oct 1 08:50:00 2020

@author: nicholasfinan
"""
import os ,shutil
from openpyxl import load_workbook
import pandas as pd
# Import Auxillary functions
import "Mqapping_Functions.py" as map

import os ,shutil
from openpyxl import load_workbook
import pandas as pd

original_path = os.getcwd()

#%% Load spreadsheet with data for all samples

xl = pd.ExcelFile(original_path+'/Samples_ma302281b.xlsx')
df  = xl.parse(sheet_name=0)
xl = pd.ExcelFile(original_path+'/Reference_Tg_values.xlsx')
df2  = xl.parse(sheet_name=0)


def mapping():    #Fill in every cell for every sample
    map.create_dir('SUBMISSION')
    nb_sample = df['sample no.'].shape[0]
    for i in range(nb_sample):
        file = original_path+'/master_template_ma302281b.xlsx'
        workbook = load_workbook(filename=file)

        map.map_data_origin(workbook ,i)
        map.map_material_type(workbook ,i)
        map.map_synthesis_process(workbook ,i)
        map.map_Tg(workbook ,i)
        map.map_microstructre(workbook, i)

        sample = 'S'+str(i+1)
        os.mkdir(sample)
        map.copy_image(i)
        workbook.save(filename=sample+'/S'+str(i+1)+'_template.xlsx')


#%% Launch program
mapping()
